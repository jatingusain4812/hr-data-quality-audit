"""
HR Data Quality Audit & Risk Control Engine
============================================
Performs automated audit checks on HR data, assigns risk scores,
generates a clean dataset, and exports results to Excel.

Tools: Python, Pandas, OpenPyXL
"""

import pandas as pd
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import re
import os
from datetime import date

# ── 1. LOAD DATA ──────────────────────────────────────────────────────────────
df = pd.read_csv("data/hr_raw_data.csv", parse_dates=["Date_of_Birth", "Join_Date", "Contract_Expiry"])
print(f"📂 Loaded {len(df)} records")

# ── 2. AUDIT CHECKS ───────────────────────────────────────────────────────────
issues = pd.DataFrame({"Employee_ID": df["Employee_ID"]})

# Check 1: Missing Email
issues["Missing_Email"] = df["Email"].isna() | df["Email"].eq("")

# Check 2: Invalid Email Format
def valid_email(e):
    if not e or pd.isna(e): return False
    return bool(re.match(r"^[\w\.\+\-]+@[\w\-]+\.[a-zA-Z]{2,}$", str(e)))

issues["Invalid_Email"] = df["Email"].apply(lambda x: not valid_email(x) and str(x).strip() != "")

# Check 3: Missing / Negative Salary
issues["Missing_Salary"]  = df["Salary"].isna()
issues["Negative_Salary"] = df["Salary"].fillna(0) < 0

# Check 4: Missing Department
issues["Missing_Dept"] = df["Department"].isna() | df["Department"].eq("")

# Check 5: Underage Employee (Risk Flag)
issues["Underage_Risk"] = df["Age"] < 18

# Check 6: Missing DOB
issues["Missing_DOB"] = df["Date_of_Birth"].isna()

# Check 7: Missing Phone
issues["Missing_Phone"] = df["Phone"].isna() | df["Phone"].eq("")

# Check 8: Missing PAN Number
issues["Missing_PAN"] = df["PAN_Number"].isna() | df["PAN_Number"].eq("")

# Check 9: Salary outlier (> 3 std deviations)
sal = df["Salary"].dropna()
mean_s, std_s = sal.mean(), sal.std()
issues["Salary_Outlier"] = df["Salary"].apply(
    lambda x: abs(x - mean_s) > 3 * std_s if pd.notna(x) and x > 0 else False
)

# Check 10: Duplicate emails
email_counts = df["Email"].value_counts()
issues["Duplicate_Email"] = df["Email"].map(email_counts) > 1

# ── 3. RISK SCORE ─────────────────────────────────────────────────────────────
weights = {
    "Underage_Risk": 5,
    "Negative_Salary": 4,
    "Missing_Salary": 3,
    "Missing_Dept": 2,
    "Invalid_Email": 2,
    "Duplicate_Email": 2,
    "Missing_PAN": 1,
    "Missing_Email": 1,
    "Missing_DOB": 1,
    "Missing_Phone": 1,
    "Salary_Outlier": 2
}

bool_cols = [c for c in weights if c in issues.columns]
issues["Risk_Score"] = sum(issues[c].astype(int) * w for c, w in weights.items() if c in issues.columns)

def risk_level(score):
    if score >= 7: return "HIGH"
    elif score >= 3: return "MEDIUM"
    elif score > 0: return "LOW"
    return "CLEAN"

issues["Risk_Level"] = issues["Risk_Score"].apply(risk_level)
issues["Total_Issues"] = issues[bool_cols].sum(axis=1)

# Merge back
audit_df = df.merge(issues.drop(columns=["Employee_ID"]), left_index=True, right_index=True)

# ── 4. SUMMARY STATS ──────────────────────────────────────────────────────────
summary = {
    "Total Employees": len(df),
    "Clean Records": (issues["Risk_Level"] == "CLEAN").sum(),
    "Low Risk": (issues["Risk_Level"] == "LOW").sum(),
    "Medium Risk": (issues["Risk_Level"] == "MEDIUM").sum(),
    "High Risk": (issues["Risk_Level"] == "HIGH").sum(),
    "Total Issues Found": issues[bool_cols].sum().sum(),
}

dept_summary = audit_df[audit_df["Department"].ne("") & audit_df["Department"].notna()].groupby("Department").agg(
    Total_Employees=("Employee_ID","count"),
    High_Risk=("Risk_Level", lambda x: (x == "HIGH").sum()),
    Avg_Risk_Score=("Risk_Score","mean"),
    Avg_Salary=("Salary","mean")
).round(2).reset_index()

issue_counts = issues[bool_cols].sum().reset_index()
issue_counts.columns = ["Issue_Type", "Count"]
issue_counts = issue_counts.sort_values("Count", ascending=False)

print("\n📊 Audit Summary:")
for k,v in summary.items():
    print(f"   {k}: {v}")

# ── 5. BUILD EXCEL REPORT ─────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# --- Color palette ---
CLR_HEADER    = "1F3864"   # dark navy
CLR_SUBHDR    = "2E75B6"   # blue
CLR_HIGH      = "C00000"   # red
CLR_MEDIUM    = "ED7D31"   # orange
CLR_LOW       = "FFD700"   # yellow
CLR_CLEAN     = "70AD47"   # green
CLR_ACCENT    = "E9F0FA"   # light blue bg
CLR_WHITE     = "FFFFFF"
CLR_ALT       = "F2F7FF"

thin = Border(
    left=Side(style='thin', color="BBBBBB"),
    right=Side(style='thin', color="BBBBBB"),
    top=Side(style='thin', color="BBBBBB"),
    bottom=Side(style='thin', color="BBBBBB")
)

def header_style(cell, bg=CLR_HEADER, fg=CLR_WHITE, size=11, bold=True):
    cell.font = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin

def data_style(cell, bg=CLR_WHITE, bold=False, color="000000", center=False):
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border = thin

def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ════════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("📊 Dashboard")
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:J2")
ws1["A1"] = "🔍  HR DATA QUALITY AUDIT — RISK CONTROL DASHBOARD"
ws1["A1"].font = Font(bold=True, size=18, color=CLR_WHITE, name="Arial")
ws1["A1"].fill = PatternFill("solid", start_color=CLR_HEADER)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 35
ws1.row_dimensions[2].height = 20

# Subtitle
ws1.merge_cells("A3:J3")
ws1["A3"] = f"Generated: {date.today()}  |  Total Records Audited: {len(df)}  |  Tool: Python + MySQL + Excel"
ws1["A3"].font = Font(italic=True, size=10, color="555555", name="Arial")
ws1["A3"].alignment = Alignment(horizontal="center")
ws1["A3"].fill = PatternFill("solid", start_color=CLR_ACCENT)

# KPI Cards row
kpi_data = [
    ("Total Employees", summary["Total Employees"], CLR_SUBHDR),
    ("✅ Clean Records", summary["Clean Records"], "70AD47"),
    ("⚠️ Medium Risk",  summary["Medium Risk"],   "ED7D31"),
    ("🚨 High Risk",    summary["High Risk"],      "C00000"),
    ("Issues Found",    int(summary["Total Issues Found"]), "7030A0"),
]

ws1.row_dimensions[5].height = 20
ws1.row_dimensions[6].height = 50
ws1.row_dimensions[7].height = 30
ws1.row_dimensions[8].height = 20

kpi_cols = [1, 3, 5, 7, 9]
for idx, (label, value, color) in enumerate(kpi_data):
    c = kpi_cols[idx]
    col_letter = get_column_letter(c)
    next_col = get_column_letter(c+1)
    ws1.merge_cells(f"{col_letter}5:{next_col}5")
    ws1.merge_cells(f"{col_letter}6:{next_col}6")
    ws1.merge_cells(f"{col_letter}7:{next_col}7")
    ws1[f"{col_letter}5"] = label
    ws1[f"{col_letter}5"].font = Font(bold=True, size=10, color=CLR_WHITE, name="Arial")
    ws1[f"{col_letter}5"].fill = PatternFill("solid", start_color=color)
    ws1[f"{col_letter}5"].alignment = Alignment(horizontal="center", vertical="center")
    ws1[f"{col_letter}6"] = value
    ws1[f"{col_letter}6"].font = Font(bold=True, size=24, color=color, name="Arial")
    ws1[f"{col_letter}6"].fill = PatternFill("solid", start_color="F8F8F8")
    ws1[f"{col_letter}6"].alignment = Alignment(horizontal="center", vertical="center")
    ws1[f"{col_letter}6"].border = thin

# Issue Breakdown Table
row = 10
ws1.merge_cells(f"A{row}:D{row}")
ws1[f"A{row}"] = "Issue Type Breakdown"
header_style(ws1[f"A{row}"], bg=CLR_SUBHDR)
ws1.row_dimensions[row].height = 22

row += 1
for col, h in zip("ABCD", ["#", "Issue Type", "Count", "% of Total"]):
    ws1[f"{col}{row}"] = h
    header_style(ws1[f"{col}{row}"], bg="2E75B6")
ws1.row_dimensions[row].height = 18

total_issues = int(issue_counts["Count"].sum())
for i, (_, r) in enumerate(issue_counts.iterrows()):
    row += 1
    bg = CLR_ALT if i % 2 == 0 else CLR_WHITE
    pct = f"{r['Count']/len(df)*100:.1f}%"
    for col, val in zip("ABCD", [i+1, r["Issue_Type"].replace("_", " "), int(r["Count"]), pct]):
        ws1[f"{col}{row}"] = val
        data_style(ws1[f"{col}{row}"], bg=bg, center=(col in "ACD"))

# Department Risk Table
row += 2
ws1.merge_cells(f"F10:J10")
ws1[f"F10"] = "Department Risk Summary"
header_style(ws1[f"F10"], bg=CLR_SUBHDR)

row2 = 11
for col, h in zip("FGHIJ", ["Department", "Employees", "High Risk", "Avg Score", "Avg Salary (₹)"]):
    ws1[f"{col}{row2}"] = h
    header_style(ws1[f"{col}{row2}"], bg="2E75B6")

for i, (_, r) in enumerate(dept_summary.iterrows()):
    row2 += 1
    bg = CLR_ALT if i % 2 == 0 else CLR_WHITE
    for col, val in zip("FGHIJ", [r["Department"], int(r["Total_Employees"]),
                                   int(r["High_Risk"]), round(r["Avg_Risk_Score"],1),
                                   f"₹{int(r['Avg_Salary']):,}"]):
        ws1[f"{col}{row2}"] = val
        data_style(ws1[f"{col}{row2}"], bg=bg, center=(col != "F"))

set_col_widths(ws1, {"A":4,"B":28,"C":10,"D":12,"E":2,"F":16,"G":12,"H":10,"I":10,"J":16})

# ════════════════════════════════════════════════════════════════
# SHEET 2: FULL AUDIT LOG
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📋 Audit Log")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "C2"

display_cols = ["Employee_ID","Full_Name","Department","Job_Role","Salary",
                "Employment_Status","Risk_Score","Risk_Level","Total_Issues"] + bool_cols

headers = display_cols
for ci, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=ci, value=h.replace("_"," "))
    header_style(cell)

ws2.row_dimensions[1].height = 30

risk_colors = {"HIGH": CLR_HIGH, "MEDIUM": "ED7D31", "LOW": "FFD700", "CLEAN": CLR_CLEAN}

for ri, (_, row_data) in enumerate(audit_df[display_cols].iterrows(), 2):
    bg = CLR_ALT if ri % 2 == 0 else CLR_WHITE
    risk = row_data["Risk_Level"]
    for ci, (col, val) in enumerate(row_data.items(), 1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        if col == "Risk_Level":
            data_style(cell, bg=risk_colors.get(risk, CLR_WHITE), bold=True, center=True)
            cell.font = Font(bold=True, color=CLR_WHITE, name="Arial", size=10)
        elif col in bool_cols and val:
            data_style(cell, bg="FFE0E0", center=True)
            cell.value = "⚠ YES"
            cell.font = Font(bold=True, color=CLR_HIGH, name="Arial", size=9)
        elif col in bool_cols:
            data_style(cell, bg=bg, center=True)
            cell.value = "✓"
            cell.font = Font(color="70AD47", name="Arial", size=10)
        elif col == "Risk_Score":
            data_style(cell, bg=bg, center=True, bold=True)
        else:
            data_style(cell, bg=bg)

col_widths = {"A":10,"B":18,"C":14,"D":20,"E":14,"F":16,"G":10,"H":10,"I":12}
for k,v in col_widths.items():
    ws2.column_dimensions[k].width = v
for c in range(9, len(display_cols)+1):
    ws2.column_dimensions[get_column_letter(c)].width = 16

# ════════════════════════════════════════════════════════════════
# SHEET 3: HIGH RISK EMPLOYEES
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🚨 High Risk")
ws3.sheet_view.showGridLines = False

high_risk = audit_df[audit_df["Risk_Level"] == "HIGH"].sort_values("Risk_Score", ascending=False)

ws3.merge_cells("A1:I1")
ws3["A1"] = f"🚨 HIGH RISK EMPLOYEES — Immediate Action Required  ({len(high_risk)} employees)"
ws3["A1"].font = Font(bold=True, size=13, color=CLR_WHITE, name="Arial")
ws3["A1"].fill = PatternFill("solid", start_color=CLR_HIGH)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

hr_cols = ["Employee_ID","Full_Name","Age","Department","Salary","Risk_Score","Risk_Level",
           "Underage_Risk","Negative_Salary","Missing_Dept"]
for ci, h in enumerate(hr_cols, 1):
    cell = ws3.cell(row=2, column=ci, value=h.replace("_"," "))
    header_style(cell, bg=CLR_HIGH)

for ri, (_, row_data) in enumerate(high_risk[hr_cols].iterrows(), 3):
    for ci, (col, val) in enumerate(row_data.items(), 1):
        cell = ws3.cell(row=ri, column=ci, value=val)
        if col in ["Underage_Risk","Negative_Salary","Missing_Dept"] and val:
            data_style(cell, bg="FFE0E0", center=True)
            cell.value = "🔴 FLAG"
            cell.font = Font(bold=True, color=CLR_HIGH, name="Arial", size=9)
        elif col == "Risk_Score":
            data_style(cell, bg="FFF0F0", bold=True, center=True)
        else:
            data_style(cell, bg="FFF8F8")

for ci in range(1, len(hr_cols)+1):
    ws3.column_dimensions[get_column_letter(ci)].width = 16

# ════════════════════════════════════════════════════════════════
# SHEET 4: CLEANED DATA
# ════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("✅ Clean Data")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "A2"

clean_df = df[issues["Risk_Level"] == "CLEAN"].copy()
clean_cols = ["Employee_ID","Full_Name","Age","Gender","Email","Phone",
              "Department","Job_Role","Join_Date","Salary","Employment_Status","PAN_Number"]

ws4.merge_cells(f"A1:{get_column_letter(len(clean_cols))}1")
ws4["A1"] = f"✅ CLEAN RECORDS — Passed All Quality Checks  ({len(clean_df)} employees)"
ws4["A1"].font = Font(bold=True, size=12, color=CLR_WHITE, name="Arial")
ws4["A1"].fill = PatternFill("solid", start_color="375623")
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 25

for ci, h in enumerate(clean_cols, 1):
    cell = ws4.cell(row=2, column=ci, value=h.replace("_"," "))
    header_style(cell, bg="375623")

for ri, (_, row_data) in enumerate(clean_df[clean_cols].iterrows(), 3):
    bg = CLR_ALT if ri % 2 == 0 else CLR_WHITE
    for ci, val in enumerate(row_data.values, 1):
        cell = ws4.cell(row=ri, column=ci, value=str(val) if pd.notna(val) else "")
        data_style(cell, bg=bg)

for ci in range(1, len(clean_cols)+1):
    ws4.column_dimensions[get_column_letter(ci)].width = 18

# ════════════════════════════════════════════════════════════════
# SHEET 5: SQL REFERENCE
# ════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🗄 SQL Queries")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:C1")
ws5["A1"] = "SQL Queries Used in This Audit (MySQL)"
ws5["A1"].font = Font(bold=True, size=13, color=CLR_WHITE, name="Arial")
ws5["A1"].fill = PatternFill("solid", start_color=CLR_HEADER)
ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 28

queries = [
    ("1. Count Records per Department",
     "SELECT department, COUNT(*) AS total_employees\nFROM hr_employees\nGROUP BY department\nORDER BY total_employees DESC;"),
    ("2. Find Missing Email Records",
     "SELECT employee_id, full_name, department\nFROM hr_employees\nWHERE email IS NULL OR email = '';"),
    ("3. Detect Negative or NULL Salaries",
     "SELECT employee_id, full_name, salary\nFROM hr_employees\nWHERE salary IS NULL OR salary < 0;"),
    ("4. Identify Underage Employees (Risk!)",
     "SELECT employee_id, full_name, age, department\nFROM hr_employees\nWHERE age < 18;"),
    ("5. High Risk Summary by Department",
     "SELECT department,\n  COUNT(*) AS total,\n  SUM(CASE WHEN risk_level='HIGH' THEN 1 ELSE 0 END) AS high_risk,\n  ROUND(AVG(risk_score),2) AS avg_risk\nFROM hr_audit\nGROUP BY department;"),
    ("6. Duplicate Email Detection",
     "SELECT email, COUNT(*) AS dup_count\nFROM hr_employees\nWHERE email != ''\nGROUP BY email\nHAVING COUNT(*) > 1;"),
    ("7. Missing PAN Numbers",
     "SELECT employee_id, full_name\nFROM hr_employees\nWHERE pan_number IS NULL OR pan_number = '';"),
    ("8. Salary Outlier Detection (>2x Avg)",
     "SELECT employee_id, full_name, salary,\n  (SELECT AVG(salary) FROM hr_employees WHERE salary > 0) AS avg_salary\nFROM hr_employees\nWHERE salary > 2 * (SELECT AVG(salary) FROM hr_employees WHERE salary > 0);"),
]

row = 3
for title, query in queries:
    ws5.merge_cells(f"A{row}:C{row}")
    ws5[f"A{row}"] = title
    ws5[f"A{row}"].font = Font(bold=True, size=11, color=CLR_WHITE, name="Arial")
    ws5[f"A{row}"].fill = PatternFill("solid", start_color=CLR_SUBHDR)
    ws5[f"A{row}"].alignment = Alignment(vertical="center")
    ws5.row_dimensions[row].height = 20
    row += 1
    line_count = query.count("\n") + 1
    ws5.merge_cells(f"A{row}:C{row + line_count}")
    ws5[f"A{row}"] = query
    ws5[f"A{row}"].font = Font(name="Courier New", size=10, color="1F3864")
    ws5[f"A{row}"].fill = PatternFill("solid", start_color="F0F4FF")
    ws5[f"A{row}"].alignment = Alignment(vertical="top", wrap_text=True)
    ws5.row_dimensions[row].height = 14 * line_count
    row += line_count + 2

ws5.column_dimensions["A"].width = 90
ws5.column_dimensions["B"].width = 1
ws5.column_dimensions["C"].width = 1

# ── SAVE ──────────────────────────────────────────────────────────────────────
os.makedirs("excel", exist_ok=True)
out_path = "excel/HR_Audit_Risk_Report.xlsx"
wb.save(out_path)
print(f"\n✅ Excel report saved → {out_path}")
print(f"   Sheets: Dashboard | Audit Log | High Risk | Clean Data | SQL Queries")
print(f"\n🎯 Project ready for GitHub!")
